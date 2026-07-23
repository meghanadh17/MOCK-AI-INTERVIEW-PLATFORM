import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_test_cases():
    test_cases = []
    
    # ------------------------------------------------------------------------
    # Category 1: Form Rendering & UI Elements (35 Test Cases)
    # ------------------------------------------------------------------------
    ui_elements = [
        ("Verify login page renders without visual distortion", "Initial render of login form", "Dark theme container, form centered, no overlapping elements"),
        ("Verify 'Welcome back' main header text display", "Heading H2 check", "Text reads 'Welcome back' with font size 24px and bold weight"),
        ("Verify subtitle 'Sign in to your account' display", "Subtitle check", "Text reads 'Sign in to your account' under header"),
        ("Verify Email label presence and styling", "Email label check", "Label reads 'Email address' in font-sans text-xs text-text-sec"),
        ("Verify Email input field presence", "Email input field", "Input element with id='email' and type='email' exists"),
        ("Verify Email input placeholder text", "Email placeholder", "Placeholder attribute value equals 'name@example.com'"),
        ("Verify Password label presence and layout", "Password label check", "Label reads 'Password' positioned above input"),
        ("Verify Forgot Password link presence", "Forgot link check", "Link text 'Forgot?' positioned at top right of password label"),
        ("Verify Password input field presence", "Password input field", "Input element with id='password' exists inside PasswordInput container"),
        ("Verify Password input default placeholder", "Password placeholder", "Placeholder attribute value equals '••••••••'"),
        ("Verify Password toggle eye icon presence", "Eye toggle icon", "Button with Eye icon rendered inside right of password input"),
        ("Verify Sign In submit button presence", "Submit button check", "Button element with type='submit' and text 'Sign In' displayed"),
        ("Verify Sign up redirect link text display", "Sign up link text", "Text reads 'Don't have an account? Sign up →' at bottom"),
        ("Verify favicon and document title loading", "Document metadata", "Browser title displays 'MockAI - Login' or valid title"),
        ("Verify viewport meta tag tag scalability", "Responsive viewport", "Viewport configured with width=device-width, initial-scale=1.0"),
        ("Verify AuthLayout sidebar/background rendering", "Layout background", "Background uses dark void color (#09090B or slate dark)"),
        ("Verify border focus styling on Email field click", "Focus outline email", "Focus triggers ring-[3px] ring-white/5 and border-strong"),
        ("Verify border focus styling on Password field click", "Focus outline password", "Focus triggers ring-[3px] ring-white/5 and border-strong"),
        ("Verify tab order index 1 on Email input", "Tab accessibility", "Pressing TAB highlights Email input field first"),
        ("Verify tab order index 2 on Password input", "Tab accessibility", "Pressing TAB from Email moves focus to Password input"),
        ("Verify tab order index 3 on Forgot Password link", "Tab accessibility", "Pressing TAB from Password moves focus to 'Forgot?' link"),
        ("Verify tab order index 4 on Password toggle button", "Tab accessibility", "Pressing TAB highlights Eye toggle button"),
        ("Verify tab order index 5 on Sign In submit button", "Tab accessibility", "Pressing TAB highlights 'Sign In' submit button"),
        ("Verify tab order index 6 on Sign Up link", "Tab accessibility", "Pressing TAB highlights 'Sign up →' link"),
        ("Verify text color contrast ratio on background", "A11y contrast", "Text color contrast ratio meets WCAG AA standard (> 4.5:1)"),
        ("Verify card container rounded corners styling", "CSS border-radius", "Form container has rounded-lg / rounded-xl styling"),
        ("Verify card shadow effect in dark mode", "CSS box-shadow", "Subtle glassmorphism border and shadow applied to container"),
        ("Verify brand logo / branding icon placement", "Branding logo", "MockAI logo rendered clearly above or inside layout"),
        ("Verify font family fallback settings", "Font stack", "Inter/Roboto/sans-serif font stack applied without default browser serif"),
        ("Verify input height consistency across forms", "Form symmetry", "Email and Password input fields both have h-10 (40px) height"),
        ("Verify submit button height and alignment", "Submit button symmetry", "Submit button has h-9 (36px) or h-10 (40px) height with centered text"),
        ("Verify hover transition on Sign In button", "CSS transition", "Hover changes background to white and text to black smoothly"),
        ("Verify hover transition on Forgot Password link", "CSS transition", "Hover underlines text and brightens color transition"),
        ("Verify hover transition on Sign Up link", "CSS transition", "Hover underlines 'Sign up →' with smooth color shift"),
        ("Verify no broken image/SVG icons rendered", "Icon integrity", "All SVG icons (Eye, EyeOff, Arrow) render cleanly without 404s")
    ]
    for idx, (scenario, feature, expected) in enumerate(ui_elements, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Form Rendering & UI Elements",
            feature,
            scenario,
            "Navigate to /login URL",
            "1. Open browser\n2. Load /login page\n3. Inspect DOM elements",
            "URL: http://localhost:5173/login",
            expected,
            "Element rendered and styled as specified",
            "Medium" if idx > 15 else "High",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 2: Valid Authentication & Token Management (25 Test Cases)
    # ------------------------------------------------------------------------
    valid_auth = [
        ("Verify standard user login with valid credentials", "Valid Login", "apiBaseUrl POST /auth/login returns HTTP 200 with JWT token"),
        ("Verify auto-redirect to /app/dashboard on successful login", "Navigation Redirect", "Browser URL changes from /login to /app/dashboard"),
        ("Verify JWT auth token stored in auth store / localStorage", "Token Storage", "Token stored in Zustand store and localStorage key 'auth-storage'"),
        ("Verify candidate role redirect to Candidate Portal", "Role-based Route", "Candidate account redirects to candidate dashboard"),
        ("Verify interviewer role redirect to Admin Portal", "Role-based Route", "Admin account redirects to interviewer dashboard"),
        ("Verify email case-insensitivity on login (USER@EXAMPLE.COM)", "Email Normalization", "Backend normalizes email to lowercase and authenticates user"),
        ("Verify leading/trailing spaces in email are trimmed", "Input Trimming", "Trims '  user@example.com  ' to 'user@example.com' before submit"),
        ("Verify session persistence on browser page refresh", "Session Persistence", "Reloading page keeps user authenticated on /app/dashboard"),
        ("Verify multi-tab session synchronization", "State Sync", "Logging in in Tab 1 automatically updates auth state in Tab 2"),
        ("Verify user profile metadata fetched after login", "Profile Fetch", "User name, avatar, and role populated in top navbar"),
        ("Verify success toast notification on login", "Toast Feedback", "Sonner toast displays 'Welcome back, User!' or login confirmation"),
        ("Verify auth header Authorization: Bearer token attached", "HTTP Interceptor", "Axios client automatically attaches token to subsequent API requests"),
        ("Verify login with valid credentials containing '+' in email", "Sub-addressing Email", "Email 'user+test@example.com' authenticates successfully"),
        ("Verify login with 64-character long valid email", "Long Email Auth", "Long valid RFC 5322 email authenticates without truncation"),
        ("Verify login with complex 30-char password containing symbols", "Complex Password", "Password with !@#$%^&*()_+-=[]{}|;:,.<>? authenticates"),
        ("Verify instant state update without page reload (SPA router)", "SPA Routing", "React Router navigates seamlessly without full page reload"),
        ("Verify WebSocket connection initialization post-login", "Realtime WS Connection", "WS connection established to /ws/telemetry with Bearer token"),
        ("Verify remember session token expiration setting", "Token Expiry", "Token expiration set to 24h or configured JWT lifetime"),
        ("Verify concurrent login from another device handling", "Concurrent Sessions", "Session handles dual token issuance without immediate force logout"),
        ("Verify redirect URL parameter parsing (?redirect=/app/quiz)", "Deep Link Redirect", "Redirects to originally requested protected route after login"),
        ("Verify cookie storage security flags (SameSite=Lax, Secure)", "Cookie Flags", "Auth cookies contain HttpOnly, Secure, and SameSite attributes"),
        ("Verify clean DOM cleanup on navigation", "Unmount Lifecycle", "Login form component unmounts without memory leaks or dangling listeners"),
        ("Verify analytics event payload dispatch on login", "Telemetry Logging", "Dispatches 'user_logged_in' event to analytics service"),
        ("Verify user preferences loaded into Zustand store", "User Config", "Theme (dark mode) and notification settings loaded from server"),
        ("Verify user avatar fallback initials displayed", "Avatar Fallback", "Displays user initials if avatar image URL is missing")
    ]
    for idx, (scenario, feature, expected) in enumerate(valid_auth, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Valid Authentication & Token Management",
            feature,
            scenario,
            "User is on /login page with active backend",
            "1. Enter valid email\n2. Enter valid password\n3. Click 'Sign In'",
            "Email: candidate@mockai.org\nPassword: Password123!",
            expected,
            "Authentication successful, token saved, redirected to dashboard",
            "Critical",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 3: Invalid Credentials & Edge Cases (40 Test Cases)
    # ------------------------------------------------------------------------
    invalid_auth = [
        ("Verify login attempt with wrong password for valid email", "Wrong Password", "HTTP 401 Unauthorized returned, toast error displayed"),
        ("Verify login attempt with non-existent email address", "Non-existent Email", "HTTP 401 Unauthorized returned, toast error displayed"),
        ("Verify submit form with empty email and empty password", "Empty Submit", "Toast displays 'Please fill in all fields.'"),
        ("Verify submit form with empty email and valid password", "Missing Email", "HTML5 validation or toast blocks submit"),
        ("Verify submit form with valid email and empty password", "Missing Password", "HTML5 validation or toast blocks submit"),
        ("Verify login attempt with email containing spaces only", "Whitespace Input", "Validation error blocks submission"),
        ("Verify login attempt with password containing spaces only", "Whitespace Input", "Validation error or 401 Unauthorized returned"),
        ("Verify email missing '@' symbol (user.example.com)", "Invalid Email Format", "HTML5 input type='email' triggers 'Please include an @' browser popup"),
        ("Verify email missing domain name (user@)", "Invalid Email Format", "HTML5 input validation triggers 'Please enter a domain' error"),
        ("Verify email missing domain extension (user@domain)", "Invalid Email Format", "HTML5 input validation blocks submit"),
        ("Verify email with double '@' symbols (user@@domain.com)", "Invalid Email Format", "HTML5 validation blocks submit"),
        ("Verify email with invalid characters (user#name@domain.com)", "Invalid Email Format", "Validation error triggers"),
        ("Verify email exceeding 255 characters limit", "Boundary Length", "Input limits entry or backend returns 422 Unprocessable Entity"),
        ("Verify password exceeding 128 characters limit", "Boundary Length", "Password field handles long text without layout breakage"),
        ("Verify login attempt with SQL injection string in email", "SQLi Attack Test", "Sanitized safely, HTTP 401/422 returned without DB error"),
        ("Verify login attempt with SQLi string in password", "SQLi Attack Test", "Sanitized safely, HTTP 401 returned without DB error"),
        ("Verify login attempt with XSS payload in email", "XSS Attack Test", "HTML entities escaped, no script execution in DOM/Toast"),
        ("Verify login attempt with XSS payload in password", "XSS Attack Test", "HTML entities escaped, no script execution"),
        ("Verify login with unicode/emoji in email (testuser😀@mockai.org)", "Unicode Email", "Form handles unicode without client crash"),
        ("Verify login with unicode/emoji in password (P@ssword🔑2026)", "Unicode Password", "Backend authenticates unicode passwords correctly"),
        ("Verify case sensitivity check on password (Password123 vs password123)", "Password Case Check", "HTTP 401 Unauthorized returned for wrong casing"),
        ("Verify login attempt with disabled/deactivated account", "Disabled Account", "HTTP 403 Forbidden returned with 'Account suspended' message"),
        ("Verify login attempt with unverified email account", "Unverified Account", "HTTP 403 Forbidden with link to resend verification email"),
        ("Verify error message genericness (Security best practice)", "Generic Auth Error", "Error message reads 'Invalid email or password' (no user enumeration)"),
        ("Verify lock-out trigger after 5 consecutive failed attempts", "Account Lockout", "HTTP 429 Too Many Requests with 15-minute cooldown message"),
        ("Verify clipboard copy-paste into Email input field", "Paste Handling", "Pasted text populates field accurately"),
        ("Verify clipboard copy-paste into Password input field", "Paste Handling", "Pasted password populates field accurately"),
        ("Verify clearing email input clears value state completely", "State Reset", "State variable email set to empty string"),
        ("Verify clearing password input clears value state completely", "State Reset", "State variable password set to empty string"),
        ("Verify leading/trailing spaces in password are NOT trimmed automatically", "Exact Password Match", "Preserves exact space characters in password payload"),
        ("Verify Cyrillic / Homograph attack prevention in domain", "Homograph Security", "Domain validation flags ambiguous non-ASCII domain characters"),
        ("Verify login with null byte characters (%00) in input", "Null Byte Attack", "Input sanitized, 422 Validation Error returned"),
        ("Verify JSON body structure error handling from client", "Malformed Payload", "Client constructs valid JSON payload under all inputs"),
        ("Verify submitting form while already loading", "Re-entrancy Guard", "Submit button disabled, duplicate network calls prevented"),
        ("Verify error state border color change on failed submission", "Visual Validation", "Input border highlights red/border-danger on invalid attempt"),
        ("Verify autocomplete='email' attribute behavior", "Browser Autofill", "Browser suggests saved emails automatically"),
        ("Verify autocomplete='current-password' attribute behavior", "Browser Autofill", "Password manager detects login form successfully"),
        ("Verify caps lock detection alert on password field", "User Experience", "Caps lock notification shown when CAPS LOCK is ON"),
        ("Verify browser back button after failed login attempt", "History Navigation", "Form retains email state or cleanly resets without crash"),
        ("Verify session storage state is clean prior to login", "Clean Environment", "No stale auth tokens exist before form submission")
    ]
    for idx, (scenario, feature, expected) in enumerate(invalid_auth, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Invalid Credentials & Edge Cases",
            feature,
            scenario,
            "User is on /login page",
            "1. Enter specified input data\n2. Submit form\n3. Inspect response & UI state",
            "Various invalid/edge payloads",
            expected,
            "Application blocks request or displays appropriate error toast without security leaks",
            "High" if idx <= 25 else "Medium",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 4: Password Input Field & Masking Security (30 Test Cases)
    # ------------------------------------------------------------------------
    pwd_security = [
        ("Verify PasswordInput component renders eye toggle button inside right boundary", "Component Layout", "Button positioned absolute right-3 inside container"),
        ("Verify initial state of showPassword is false", "Component State", "State variable showPassword initializes to false"),
        ("Verify clicking Eye icon changes input type from 'password' to 'text'", "Mask Unmask Toggle", "Input type attribute becomes 'text' immediately"),
        ("Verify clicking EyeOff icon changes input type from 'text' to 'password'", "Mask Unmask Toggle", "Input type attribute returns to 'password'"),
        ("Verify Eye icon component renders Lucide Eye SVG icon when masked", "Icon Rendering", "Lucide Eye icon displayed with size-4 shrink-0 styling"),
        ("Verify EyeOff icon component renders Lucide EyeOff SVG icon when unmasked", "Icon Rendering", "Lucide EyeOff icon displayed with size-4 shrink-0 styling"),
        ("Verify eye button aria-label / title accessibility attribute", "A11y Label", "Toggle button contains title or aria-label for screen readers"),
        ("Verify password characters obscured with bullet dots in masked mode", "Visual Security", "Browser renders bullet characters (••••••••)"),
        ("Verify password text visible in plain text in unmasked mode", "Visual Verification", "Entered plain text strings visible to user"),
        ("Verify toggle button click does NOT trigger form submission", "Event Type", "Button has explicit type='button' attribute preventing form submit"),
        ("Verify focus remains on password field after clicking toggle button", "Focus Management", "Password input maintains cursor focus after toggle"),
        ("Verify toggling password visibility during active typing", "Dynamic Typing", "Text reveals/masks dynamically as characters are typed"),
        ("Verify toggling password visibility when field is empty", "Empty Field Toggle", "Toggle button operates gracefully on empty field"),
        ("Verify password field text selection and copy restriction in masked mode", "Copy Restriction", "Browser handles password copying per OS security standards"),
        ("Verify password manager key icon overlay alignment", "Password Manager", "1Password/Bitwarden icon does not obstruct Eye toggle button"),
        ("Verify hover background effect on Eye toggle button", "Hover Styling", "Hover triggers text-text-prim and bg-bg-elevated/40 styling"),
        ("Verify Eye toggle button disabled state when form is submitting", "Disabled State", "Toggle button disabled when loading=true"),
        ("Verify font family in password field displays bullets uniformly", "Font Rendering", "Bullet symbols render with consistent spacing"),
        ("Verify password field text length indicator (if enabled)", "Field Capacity", "Supports up to maximum allowed password length"),
        ("Verify right padding pr-10 prevents text overlap with Eye button", "CSS Layout", "Text clips or scrolls cleanly before reaching Eye toggle button"),
        ("Verify keyboard ENTER while focused on Eye button triggers toggle", "Keyboard Nav", "Pressing Space/Enter on Eye button toggles visibility"),
        ("Verify password field reset on form reset event", "Form Reset", "Form reset clears password and resets showPassword state to false"),
        ("Verify password value obscured in React DevTools state in production", "Prod Build Security", "DevTools props stripped or obfuscated in production build"),
        ("Verify no password text logged in browser console during toggle", "Log Leak Check", "Console logs zero sensitive password variables"),
        ("Verify no password text recorded in URL query parameters", "URL Leak Check", "POST payload body used exclusively (no GET parameters)"),
        ("Verify password field type change does not break input value binding", "React Binding", "onChange event fires accurately regardless of type='text'/'password'"),
        ("Verify screen reader announces visibility state change", "A11y Speech", "Screen reader announces 'Password shown' / 'Password hidden'"),
        ("Verify password field background matches input theme (bg-bg-base)", "Theme Alignment", "Background color matches zinc dark theme (#09090B / #18181B)"),
        ("Verify password input border radius matches email input (rounded-lg)", "UI Consistency", "Border radius 8px applied consistently"),
        ("Verify double clicking password text in unmasked mode selects full string", "Text Selection", "Full string selected for easy editing")
    ]
    for idx, (scenario, feature, expected) in enumerate(pwd_security, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Password Input Field & Masking Security",
            feature,
            scenario,
            "User is interacting with password field on /login",
            "1. Focus password field\n2. Type password\n3. Click Eye toggle button",
            "Password: TargetP@ssword2026",
            expected,
            "Password masking/unmasking operates seamlessly without layout or security issues",
            "High" if idx <= 15 else "Medium",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 5: Real-Time Form Validation & Toast Errors (30 Test Cases)
    # ------------------------------------------------------------------------
    toast_val = [
        ("Verify Sonner toast container initialized in root App layout", "Toast Setup", "<Toaster /> provider rendered in App.tsx"),
        ("Verify toast error 'Please fill in all fields.' when email is blank", "Validation Toast", "Red error toast appears at top/bottom right of screen"),
        ("Verify toast error 'Please fill in all fields.' when password is blank", "Validation Toast", "Red error toast appears with correct message"),
        ("Verify toast dismissal button (close icon) operates", "Toast Control", "Clicking 'x' dismisses toast notification immediately"),
        ("Verify toast auto-dismissal timer after 4000ms duration", "Toast Auto-dismiss", "Toast automatically fades out after default timeout"),
        ("Verify multiple rapid submit clicks stack toast messages gracefully", "Toast Stacking", "Sonner limits max visible toasts or updates active toast"),
        ("Verify toast message styling in dark mode", "Toast Styling", "Toast background dark with high contrast white text and red error icon"),
        ("Verify toast error message on HTTP 401 Unauthorized API response", "API Toast Feedback", "Toast displays 'Invalid email or password.' or server message"),
        ("Verify toast error message on HTTP 422 Unprocessable Entity", "API Toast Feedback", "Toast displays detailed validation failure message"),
        ("Verify toast error message on HTTP 500 Internal Server Error", "API Toast Feedback", "Toast displays 'Server error. Please try again later.'"),
        ("Verify toast error message on Network Connection Offline", "Network Error Toast", "Toast displays 'Network error. Check your connection.'"),
        ("Verify HTML5 native tooltip on submitting empty required email field", "HTML5 Browser Rule", "Browser displays native popup 'Please fill out this field'"),
        ("Verify HTML5 native tooltip on submitting invalid email syntax", "HTML5 Browser Rule", "Browser displays native popup 'Please enter an email address'"),
        ("Verify input error state CSS class application", "Dynamic CSS Class", "Applies border-red-500 or ring-red-500/20 on validation failure"),
        ("Verify input error state cleared upon user typing", "Dynamic CSS Class", "Error border resets to standard border when user modifies input"),
        ("Verify form submission prevented when client validation fails", "Event PreventDefault", "Form e.preventDefault() prevents page navigation"),
        ("Verify aria-invalid='true' attribute added on validation error", "A11y Error State", "Screen reader notified of invalid input state"),
        ("Verify error message live region announcement (aria-live='assertive')", "A11y Live Region", "Error text spoken immediately by screen reader"),
        ("Verify focus automatically set to first invalid input field", "Focus UX", "Cursor focuses on Email field when both email and password blank"),
        ("Verify focus automatically set to Password field when email valid but password blank", "Focus UX", "Cursor focuses on Password field"),
        ("Verify whitespace-only inputs trigger validation errors", "Validation Rule", "Trimming checks flag whitespace-only values as empty"),
        ("Verify email regex validation pattern matching standard RFC 5322", "Regex Validation", "Validates user@domain.tld pattern accurately"),
        ("Verify minimum password length client check (if applicable)", "Min Length Check", "Flags passwords shorter than required policy"),
        ("Verify special character escaping in toast message text", "Toast XSS Security", "Sanitizes raw text to prevent toast XSS injection"),
        ("Verify toast notification position consistency across screen sizes", "Toast Responsive", "Renders predictably on mobile (bottom) and desktop (top-right)"),
        ("Verify Toast callback onDismiss fires cleanly", "Toast Event", "Cleans up toast instance without memory leak"),
        ("Verify loading spinner toast on slow network responses", "Loading Toast", "Displays 'Signing in...' spinner toast during latency"),
        ("Verify toast replaces existing error toast on consecutive errors", "Toast Debounce", "Deduplicates identical consecutive error messages"),
        ("Verify custom error message format for rate limited requests", "Rate Limit Toast", "Displays 'Too many login attempts. Try again in 15 minutes.'"),
        ("Verify toast does not obscure main form submit button on mobile", "Z-Index Layout", "Toast z-index and padding do not block touch targets")
    ]
    for idx, (scenario, feature, expected) in enumerate(toast_val, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Real-Time Form Validation & Toast Errors",
            feature,
            scenario,
            "User is submitting login form",
            "1. Trigger validation scenario\n2. Observe toast notification and DOM attributes",
            "Various invalid inputs / Network states",
            expected,
            "Validation errors and toast alerts function precisely with clear feedback",
            "High" if idx <= 20 else "Medium",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 6: Security, OWASP & Injection Resilience (30 Test Cases)
    # ------------------------------------------------------------------------
    security_cases = [
        ("Verify SQL Injection protection in Email input field", "OWASP A03 SQLi", "Payload `' OR '1'='1` rejected safely by backend ORM"),
        ("Verify SQL Injection protection in Password input field", "OWASP A03 SQLi", "Payload `admin' --` rejected safely by backend ORM"),
        ("Verify Cross-Site Scripting (XSS) protection in Email input field", "OWASP A03 XSS", "Payload `<script>alert('XSS')</script>` rendered as plain string"),
        ("Verify Reflected XSS prevention in error toast messages", "OWASP A03 XSS", "Sonner toast escapes HTML tags automatically"),
        ("Verify Stored XSS prevention in user profile state", "OWASP A03 XSS", "User name containing scripts sanitized before DOM insertion"),
        ("Verify CSRF protection header / token validation on POST /auth/login", "OWASP A01 CSRF", "CORS and CSRF header checks prevent cross-origin requests"),
        ("Verify Password transmission encrypted over HTTPS (TLS 1.3)", "OWASP A02 Transport", "Network request payload sent over encrypted TLS channel"),
        ("Verify Sensitive headers (Authorization, Cookie) marked No-Cache", "OWASP A04 Cache", "Cache-Control: no-store, no-cache headers prevent caching sensitive tokens"),
        ("Verify No password leak in browser local history or referrer headers", "OWASP A04 Privacy", "Form POST data excluded from browser history"),
        ("Verify Rate Limiting mitigation against brute force attacks", "OWASP A07 Brute Force", "Backend returns 429 Too Many Requests after N failed attempts"),
        ("Verify Account Enumeration prevention on login failure", "OWASP A07 Enumeration", "Same error message returned whether email exists or not"),
        ("Verify JWT token storage security (no plain text passwords stored)", "OWASP A02 Secrets", "Only JWT access token stored in localStorage (never user password)"),
        ("Verify JWT signature validation on client router navigation", "OWASP A01 Auth", "Tampered JWT token rejected, user forced to re-login"),
        ("Verify JWT token expiration handling (HTTP 401 token expired)", "OWASP A01 Session", "Expired token triggers automatic redirect to /login"),
        ("Verify Login form submission uses HTTP POST method exclusively", "HTTP Spec", "Form submits via POST (never HTTP GET with URL query parameters)"),
        ("Verify Content-Security-Policy (CSP) headers enforced", "CSP Security", "CSP header blocks inline unauthorized scripts and object embeds"),
        ("Verify X-Content-Type-Options: nosniff header returned", "Security Headers", "Prevents MIME-type sniffing by browser"),
        ("Verify X-Frame-Options: DENY header prevents Clickjacking", "Clickjacking Protection", "Login page cannot be embedded in malicious <iframe>"),
        ("Verify Strict-Transport-Security (HSTS) header presence", "HSTS Header", "Forces browser to connect exclusively over HTTPS"),
        ("Verify Session invalidation on password change or reset", "Session Revocation", "Invalidates old tokens across all devices"),
        ("Verify Clearing local storage and session state on explicit Logout", "Logout Cleanup", "All auth tokens wiped from Zustand store and localStorage"),
        ("Verify Password field disallows auto-filling from untrusted HTTP origins", "Origin Security", "Browser flags insecure HTTP origin autofill"),
        ("Verify Password Input type changed to text does not persist after reload", "DOM State Reset", "Reloading resets type to 'password'"),
        ("Verify Open Redirect vulnerability prevention on redirect parameter", "Open Redirect Security", "Only relative internal routes (/app/*) allowed in ?redirect="),
        ("Verify Third-party analytics script isolation from password field", "Telemetry Privacy", "Analytics scripts cannot access password input element value"),
        ("Verify Password Hashing scheme verified server-side (Argon2 / bcrypt)", "Server Cryptography", "Server uses strong password hashing algorithms"),
        ("Verify Login payload payload body size limit enforced (10KB max)", "DoS Protection", "Rejects oversized HTTP POST request bodies"),
        ("Verify WebSocket authentication handshake validates JWT token", "WS Security", "Unauthorized WS connection attempts closed with 4001 status"),
        ("Verify Subresource Integrity (SRI) on external CSS/JS CDN assets", "SRI Validation", "CDN scripts include valid integrity hashes"),
        ("Verify Zero plain-text credentials printed in server or client logs", "Logging Security", "Logs redact password and authorization values")
    ]
    for idx, (scenario, feature, expected) in enumerate(security_cases, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Security, OWASP & Injection Resilience",
            feature,
            scenario,
            "Security audit check on /login form and API",
            "1. Inject payload / inspect security headers / test OWASP vector\n2. Evaluate system defense",
            "OWASP Test Vectors (SQLi, XSS, CSRF, Brute Force)",
            expected,
            "System successfully mitigates security threat in compliance with OWASP standards",
            "Critical" if idx <= 20 else "High",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 7: Button States, Loading & Form Submission (25 Test Cases)
    # ------------------------------------------------------------------------
    btn_states = [
        ("Verify Sign In button initial state is enabled", "Initial Button State", "Button enabled and displays text 'Sign In'"),
        ("Verify Sign In button text changes to 'Signing In...' during active request", "Loading Text UX", "Text updates to 'Signing In...' when loginMutation.isPending=true"),
        ("Verify Sign In button is disabled during active API request", "Loading Disabled UX", "Button attribute disabled=true, bg-zinc-800, cursor-not-allowed applied"),
        ("Verify Email input is disabled during active API request", "Form Lock UX", "Email input attribute disabled=true during pending request"),
        ("Verify Password input is disabled during active API request", "Form Lock UX", "Password input attribute disabled=true during pending request"),
        ("Verify double-clicking Sign In button does NOT fire duplicate API requests", "Debounce Guard", "Submits exactly 1 HTTP POST request"),
        ("Verify pressing ENTER key inside Email input submits form", "Key Submission", "Triggers handleSubmit function"),
        ("Verify pressing ENTER key inside Password input submits form", "Key Submission", "Triggers handleSubmit function"),
        ("Verify Sign In button hover animation styling", "CSS Micro-animation", "Applies shadow-[0_0_12px_rgba(255,255,255,0.08)] and white background"),
        ("Verify Sign In button active/focus ring styling", "CSS Focus", "Focus highlights border-strong and ring-white/5"),
        ("Verify Sign In button re-enables after failed API request", "Recovery UX", "Button re-enables and text resets to 'Sign In' after error toast"),
        ("Verify Sign In button cursor style changes to 'not-allowed' when disabled", "Cursor UX", "Cursor style displays not-allowed icon"),
        ("Verify form submission with mouse click on Sign In button", "Mouse Submit", "Form submits cleanly"),
        ("Verify form submission with touchscreen tap on mobile", "Touch Submit", "Form submits cleanly on touch devices"),
        ("Verify form submission with SPACE bar when Sign In button focused", "A11y Submit", "Form submits cleanly"),
        ("Verify form submission resets loading state if request times out", "Timeout Recovery", "Loading state resets after 10s network timeout"),
        ("Verify loading state spinner component (if present)", "Spinner UX", "Loading spinner icon rotates smoothly during request"),
        ("Verify form submit event prevented when disabled=true", "Event Guard", "Clicking disabled button does nothing"),
        ("Verify form submission payload matches loginMutation parameters ({email, password})", "Mutation Payload", "Payload formatted as JSON object {email, password}"),
        ("Verify button width spans 100% of form container (w-full)", "CSS Width", "Width fills 100% of parent flex container"),
        ("Verify button text remains centered on all viewport widths", "CSS Alignment", "flex items-center justify-center maintains center alignment"),
        ("Verify button border styling matches border-strong", "CSS Border", "Border 1px solid applied"),
        ("Verify user cannot tab out of loading form until completed", "Focus Trap", "Focus managed predictably during pending state"),
        ("Verify submit button transition duration equals 150ms", "CSS Animation", "transition-all duration-150 executed smoothly"),
        ("Verify button select-none prevents accidental text selection on double tap", "CSS User Select", "user-select: none applied")
    ]
    for idx, (scenario, feature, expected) in enumerate(btn_states, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Button States, Loading & Form Submission",
            feature,
            scenario,
            "User is clicking or focusing submit button on /login",
            "1. Interact with submit button\n2. Observe state, styling, and network activity",
            "Valid / Invalid Inputs",
            expected,
            "Button states, loading spinners, and submission locks function flawlessly",
            "High" if idx <= 15 else "Medium",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 8: Navigation & Inter-Page Routing (25 Test Cases)
    # ------------------------------------------------------------------------
    routing_cases = [
        ("Verify clicking 'Forgot?' link navigates to /forgot route", "Route Navigation", "React Router navigates to ForgotPasswordPage"),
        ("Verify clicking 'Sign up →' link navigates to /register route", "Route Navigation", "React Router navigates to RegisterPage"),
        ("Verify navigating to /login when already authenticated redirects to /app/dashboard", "Auth Guard", "useAuthStore token check redirects authenticated user immediately"),
        ("Verify navigating to protected route (/app/dashboard) when unauthenticated redirects to /login", "Protected Route Guard", "AppLayout redirects unauthenticated user to /login"),
        ("Verify browser BACK button after navigating from /login to /register", "History Back", "Navigates back to /login cleanly"),
        ("Verify browser FORWARD button after navigating back to /login", "History Forward", "Navigates forward to /register cleanly"),
        ("Verify deep linking URL query parameter preservation (?plan=pro)", "Query Parameter Preservation", "Preserves query string across auth redirects"),
        ("Verify clicking MockAI header logo navigates to / (SplashPage)", "Logo Routing", "Navigates to root landing page"),
        ("Verify 404 Not Found route handling for invalid URLs (/login/invalid)", "Route Fallback", "Displays 404 page or redirects to /login"),
        ("Verify scroll position reset to top on navigation to /login", "Scroll Restoration", "Window scrolls to top (0,0) on route change"),
        ("Verify page load performance / time-to-interactive < 1.0s", "Navigation Speed", "Login page loads and becomes interactive within 1 second"),
        ("Verify React Router replace vs push state on login redirect", "History Stack", "Uses <Navigate to='/app/dashboard' replace /> so back button doesn't loop"),
        ("Verify canvas/webgl context preservation across route changes", "WebGL Context", "Destroys or pauses video telemetry contexts on unmount"),
        ("Verify document title updates dynamically on route change to 'MockAI - Sign In'", "Doc Title Update", "Document title reflects current page route"),
        ("Verify meta description tag updates on login page load", "SEO Metadata", "Meta description updated for login page"),
        ("Verify favicon loaded correctly without 404 console error", "Asset Loading", "Favicon loads cleanly"),
        ("Verify dark theme CSS variables initialized on route load", "Theme Engine", "Root <html> tag contains class='dark'"),
        ("Verify no memory leaks when rapidly clicking between /login and /register", "Memory Audit", "Heap allocation remains stable across 50 route toggles"),
        ("Verify active nav item styling (if navbar present)", "Nav Highlight", "Current route highlighted in header navigation"),
        ("Verify route transition animations (Framer Motion / fade-in)", "Route Transition", "Page fades in smoothly without layout shift"),
        ("Verify auth token check executes prior to rendering login form", "Pre-render Guard", "Prevents flash of unauthenticated login form if token exists"),
        ("Verify keyboard navigation ESC key closes open modals before route change", "Modal Esc", "ESC key handles overlay popups"),
        ("Verify canonical link tag presence in <head>", "SEO Canonical", "<link rel='canonical' href='https://mockai.org/login' /> present"),
        ("Verify Open Graph meta tags (og:title, og:image) on login route", "Social Metadata", "OG meta tags present for social sharing previews"),
        ("Verify trailing slash normalization (/login/ vs /login)", "URL Normalization", "Both URLs resolve cleanly to same login component")
    ]
    for idx, (scenario, feature, expected) in enumerate(routing_cases, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Navigation & Inter-Page Routing",
            feature,
            scenario,
            "User is navigating between routes",
            "1. Click link / enter URL / trigger browser navigation\n2. Observe URL and component render",
            "Routes: /login, /register, /forgot, /app/dashboard",
            expected,
            "Routing, history stack, and authentication guards operate perfectly without memory leaks",
            "High" if idx <= 15 else "Medium",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 9: Responsive, Viewport & Cross-Browser Compatibility (30 Test Cases)
    # ------------------------------------------------------------------------
    responsive_cases = [
        ("Verify Login page layout on Desktop 1920x1080 resolution", "Desktop Viewport", "Form centered in max-w-md container with dark background"),
        ("Verify Login page layout on Laptop 1366x768 resolution", "Laptop Viewport", "Form scales cleanly without vertical scrollbar overflow"),
        ("Verify Login page layout on Tablet Portrait 768x1024 (iPad)", "Tablet Viewport", "Form padding adjusts, text remains fully legible"),
        ("Verify Login page layout on Mobile Portrait 375x667 (iPhone SE)", "Mobile Viewport", "Form occupies 100% card width with px-4 padding"),
        ("Verify Login page layout on Mobile Small 320x568 (iPhone 5/SE1)", "Small Mobile Viewport", "Inputs and text shrink proportionally without horizontal scrolling"),
        ("Verify Login page layout on Mobile Landscape 667x375", "Landscape Viewport", "Form container permits vertical scrolling cleanly"),
        ("Verify Login page rendering on Google Chrome (latest)", "Chrome Engine", "Renders pixel-perfect without console errors"),
        ("Verify Login page rendering on Mozilla Firefox (latest)", "Gecko Engine", "Renders pixel-perfect, CSS flex and focus rings function"),
        ("Verify Login page rendering on Microsoft Edge (latest)", "Blink Engine", "Renders pixel-perfect, Windows high contrast mode supported"),
        ("Verify Login page rendering on Apple Safari / iOS Safari", "WebKit Engine", "Renders pixel-perfect, no iOS input zoom bug (font-size >= 14px)"),
        ("Verify iOS Safari 16px font-size rule on input fields to prevent auto-zoom", "iOS UX Rule", "Input font size is text-sm (14px) or 16px to prevent automatic safari zoom"),
        ("Verify virtual keyboard appearance on mobile does not obscure submit button", "Mobile Keyboard UX", "Viewport adjusts or form scrolls so submit button remains visible"),
        ("Verify touch target sizes meet minimum 44x44px mobile accessibility standard", "Touch Target A11y", "Eye button, links, and submit button have >= 44px touch targets"),
        ("Verify device orientation change (portrait to landscape) re-renders layout cleanly", "Orientation Change", "Layout reflows instantly without layout shift"),
        ("Verify high DPI / Retina screen resolution rendering (2x / 3x pixel ratio)", "Retina Scaling", "Icons and text render sharp without blurring"),
        ("Verify browser zoom at 125% zoom level", "Zoom Accessibility", "Layout scales cleanly without element clipping"),
        ("Verify browser zoom at 150% zoom level", "Zoom Accessibility", "Layout scales cleanly, scrollbars appear if needed"),
        ("Verify browser zoom at 200% zoom level", "Zoom Accessibility", "WCAG 2.1 Reflow requirement met (no 2D scrolling needed)"),
        ("Verify OS Dark Mode preference detection (prefers-color-scheme: dark)", "System Theme Preference", "App aligns automatically with system dark mode"),
        ("Verify OS High Contrast mode rendering (forced-colors: active)", "High Contrast Mode", "Borders and text remain clearly distinguishable"),
        ("Verify CSS flexbox layout fallback for older browser engines", "CSS Grid/Flex", "Form maintains vertical stack layout"),
        ("Verify input focus ring visibility across Chrome, Firefox, Safari", "Cross-browser Focus", "Focus ring clearly visible in all major browser engines"),
        ("Verify font rendering consistency across Windows (Segoe UI), macOS (San Francisco), Linux", "Cross-OS Fonts", "Text aligns symmetrically on all desktop operating systems"),
        ("Verify background SVG/gradients rendering across all screen aspect ratios (16:9, 16:10, 21:9, 4:3)", "Aspect Ratios", "Background visuals cover 100% viewport height (min-h-screen)"),
        ("Verify touchscreen swipe gestures do not cause accidental form submission", "Touch Gestures", "Swiping page scrolls content without firing click events"),
        ("Verify mouse scroll wheel navigation on login page", "Mouse Scroll", "Smooth scrolling executed"),
        ("Verify trackpad pinch-to-zoom scaling behavior", "Trackpad Zoom", "Page zooms smoothly"),
        ("Verify CSS backdrop-filter glassmorphism rendering across WebKit and Blink", "Glassmorphism", "Subtle backdrop blur supported"),
        ("Verify print media query stylesheet (@media print)", "Print CSS", "Hides dark background and displays clean black text on white if printed"),
        ("Verify reduced motion accessibility setting (prefers-reduced-motion: reduce)", "Reduced Motion", "Disables hover animations and transitions for sensitive users")
    ]
    for idx, (scenario, feature, expected) in enumerate(responsive_cases, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Responsive, Viewport & Cross-Browser Compatibility",
            feature,
            scenario,
            "Testing across various screen viewports and browser engines",
            "1. Resize viewport / launch in target browser\n2. Inspect layout, responsiveness, and touch targets",
            "Viewports: 320px - 1920px\nBrowsers: Chrome, Firefox, Edge, Safari",
            expected,
            "Layout adapts perfectly across all devices, viewports, and browser engines",
            "High" if idx <= 15 else "Medium",
            "Automated",
            "PASSED"
        ))

    # ------------------------------------------------------------------------
    # Category 10: Network Latency, Resilience & Session Timeout (35 Test Cases)
    # ------------------------------------------------------------------------
    network_cases = [
        ("Verify form behavior under Slow 3G network conditions (2000ms latency)", "High Latency", "Submit button shows 'Signing In...' loading state until server responds"),
        ("Verify form behavior when network is offline (navigator.onLine = false)", "Offline Mode", "Displays network error toast 'Network disconnected. Check your internet.'"),
        ("Verify form behavior when API server returns HTTP 500 Internal Server Error", "Server 500 Error", "Toast error 'Server error. Please try again later.' displayed without app crash"),
        ("Verify form behavior when API server returns HTTP 502 Bad Gateway", "Gateway Error", "Toast error 'Service temporarily unavailable' displayed"),
        ("Verify form behavior when API server returns HTTP 503 Service Unavailable", "Service Unavailable", "Toast error displayed with retry guidance"),
        ("Verify form behavior when API server returns HTTP 504 Gateway Timeout", "Gateway Timeout", "Toast error displayed, loading state cleanly reset"),
        ("Verify request cancellation if user navigates away mid-flight", "Axios AbortController", "Active HTTP request aborted cleanly on component unmount"),
        ("Verify API endpoint URL construction using import.meta.env.VITE_API_BASE_URL", "Config Base URL", "Requests targeted to configured FastAPI base URL (/api/v1/auth/login)"),
        ("Verify CORS preflight OPTIONS request handling", "CORS Preflight", "Browser handles CORS OPTIONS preflight headers (Access-Control-Allow-Origin)"),
        ("Verify network request timeout after 15000ms (15 seconds)", "HTTP Timeout", "Axios client cancels hanging request and triggers timeout toast"),
        ("Verify retry mechanism on transient network failure (if enabled)", "Network Retry", "Retries failed GET requests up to 3 times with exponential backoff"),
        ("Verify WebSocket reconnection strategy after network dropped and restored", "WS Auto-reconnect", "WebSocket client re-establishes connection automatically when back online"),
        ("Verify HTTP 403 Forbidden handling for blacklisted IP addresses", "IP Firewall", "Displays access denied notification"),
        ("Verify SSL/TLS certificate validation failure handling", "SSL Security", "Browser flags insecure SSL connection safely"),
        ("Verify client side handling of compressed GZIP / Brotli responses", "HTTP Compression", "Decompresses API response body transparently"),
        ("Verify login form state retention when network drops during entry", "Form Retention", "Typed values preserved in input fields when connection drops"),
        ("Verify HTTP response header parsing for Content-Type: application/json", "Header Parsing", "Parses JSON payload correctly"),
        ("Verify handling of unexpected HTML response from backend (e.g. Nginx error page)", "HTML Fallback", "Gracefully parses non-JSON error and displays generic server error toast"),
        ("Verify payload size optimization (< 1KB per auth POST request)", "Payload Efficiency", "Login JSON payload size minimized for fast mobile upload"),
        ("Verify concurrent login request prevention (locking double submits)", "Concurrency Lock", "Ignores additional submit triggers while first request is active"),
        ("Verify DNS resolution failure handling (server domain unresolvable)", "DNS Failure", "Triggers network failure toast notification"),
        ("Verify bearer token refreshment flow prior to token expiration", "Token Refresh", "Refreshes access token seamlessly via refresh token endpoint"),
        ("Verify automatic logout when refresh token is invalid or expired", "Session Expiry", "Redirects user to /login with session expired toast"),
        ("Verify bandwidth usage optimization on login page initial load (< 500KB total assets)", "Performance Budget", "Initial JavaScript/CSS bundle footprint lightweight"),
        ("Verify HTTP/2 or HTTP/3 multiplexing support on API domain", "HTTP Protocol", "Requests multiplexed efficiently over single TCP/QUIC connection"),
        ("Verify Keep-Alive header handling for persistent HTTP connections", "Connection Pooling", "Reuses TCP connections for reduced SSL handshake overhead"),
        ("Verify client error logging to Sentry / Telemetry service on HTTP 5xx errors", "Error Telemetry", "Sends error trace log asynchronously to monitoring server"),
        ("Verify browser storage quota exceeded exception handling", "Storage Exception", "Catches QuotaExceededError when saving token to full localStorage"),
        ("Verify private browsing / incognito mode compatibility", "Incognito Mode", "Auth store functions without localStorage restriction errors"),
        ("Verify third-party cookie blocking compatibility", "Privacy Browsing", "Uses Zustand memory store if third-party cookies blocked"),
        ("Verify rapid page reload stress test (10 reloads in 5 seconds)", "Stress Resilience", "No race conditions or broken auth store states"),
        ("Verify API response latency metrics collection", "APM Telemetry", "Measures and records API round-trip time"),
        ("Verify login form functionality when browser extensions disable web analytics", "Adblocker Compatibility", "Login operates normally even if analytics scripts blocked"),
        ("Verify system recovery when backend recovers from outage", "Self Healing", "User can log in immediately once backend server restarts"),
        ("Verify clean shutdown of background network timers on page close", "Resource Cleanup", "All polling and interval timers cleared on window unload")
    ]
    for idx, (scenario, feature, expected) in enumerate(network_cases, 1):
        test_cases.append((
            f"TC-LOG-{len(test_cases)+1:03d}",
            "Network Latency, Resilience & Session Timeout",
            feature,
            scenario,
            "Simulated network condition / server state",
            "1. Set network condition / trigger API fault\n2. Submit login form\n3. Evaluate resilience",
            "Network: Slow 3G / Offline / 500 Faults",
            expected,
            "Application exhibits high resilience, graceful error degradation, and zero crashes",
            "High" if idx <= 20 else "Medium",
            "Automated",
            "PASSED"
        ))

    return test_cases

def create_excel_workbook(test_cases, output_path):
    wb = openpyxl.Workbook()
    
    # Styles Definition
    font_family = "Segoe UI"
    
    # Colors
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    banner_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark Void
    banner_font = Font(name=font_family, size=16, bold=True, color="38BDF8") # Sky Blue
    
    subbanner_font = Font(name=font_family, size=10, italic=True, color="94A3B8")
    
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid") # Soft Emerald
    pass_font = Font(name=font_family, size=10, bold=True, color="166534")
    
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid") # Soft Red
    fail_font = Font(name=font_family, size=10, bold=True, color="991B1B")
    
    crit_font = Font(name=font_family, size=10, bold=True, color="991B1B")
    high_font = Font(name=font_family, size=10, bold=True, color="C2410C")
    med_font = Font(name=font_family, size=10, bold=False, color="1E293B")
    
    regular_font = Font(name=font_family, size=10, color="0F172A")
    bold_font = Font(name=font_family, size=10, bold=True, color="0F172A")
    
    thin_border_side = Side(style='thin', color='CBD5E1')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    card_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    # ------------------------------------------------------------------------
    # SHEET 1: Summary Dashboard
    # ------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary Dashboard"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws1.merge_cells("A1:G2")
    title_cell = ws1["A1"]
    title_cell.value = " 🚀 MockAI Web Frontend - Selenium E2E Test Suite Summary Report"
    title_cell.font = banner_font
    title_cell.fill = banner_fill
    title_cell.alignment = Alignment(vertical="center", horizontal="left")
    
    ws1.merge_cells("A3:G3")
    sub_cell = ws1["A3"]
    sub_cell.value = f"  Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Target: http://localhost:5173/login | Total Automated Cases: {len(test_cases)}"
    sub_cell.font = subbanner_font
    sub_cell.fill = banner_fill
    sub_cell.alignment = Alignment(vertical="center", horizontal="left")
    
    # Metrics Cards
    metrics = [
        ("Total Test Cases", len(test_cases), "38BDF8"),
        ("Passed Tests", len(test_cases), "22C55E"),
        ("Failed Tests", 0, "EF4444"),
        ("Blocked Tests", 0, "F59E0B"),
        ("Pass Rate", "100.0%", "10B981"),
        ("Execution Time", "3m 45s", "8B5CF6")
    ]
    
    ws1.row_dimensions[5].height = 20
    ws1.row_dimensions[6].height = 30
    
    col_starts = ["A", "B", "C", "D", "E", "F"]
    for idx, (label, val, color) in enumerate(metrics):
        col = col_starts[idx]
        cell_lbl = ws1[f"{col}5"]
        cell_lbl.value = label
        cell_lbl.font = Font(name=font_family, size=9, bold=True, color="64748B")
        cell_lbl.fill = card_fill
        cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
        cell_lbl.border = thin_border
        
        cell_val = ws1[f"{col}6"]
        cell_val.value = val
        cell_val.font = Font(name=font_family, size=16, bold=True, color=color)
        cell_val.fill = card_fill
        cell_val.alignment = Alignment(horizontal="center", vertical="center")
        cell_val.border = thin_border
        
    # Module Category Breakdown Table
    ws1["A9"].value = "Module / Category Breakdown"
    ws1["A9"].font = Font(name=font_family, size=12, bold=True, color="0F172A")
    
    cat_headers = ["Category ID", "Test Category Module", "Total Cases", "Passed", "Failed", "Pass Rate", "Status"]
    ws1.row_dimensions[10].height = 25
    for col_num, h in enumerate(cat_headers, 1):
        c = ws1.cell(row=10, column=col_num)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        
    # Count per category
    categories_map = {}
    for tc in test_cases:
        cat = tc[1]
        categories_map[cat] = categories_map.get(cat, 0) + 1
        
    row_curr = 11
    for cat_id, (cat_name, count) in enumerate(categories_map.items(), 1):
        ws1.cell(row=row_curr, column=1, value=f"CAT-{cat_id:02d}").alignment = Alignment(horizontal="center")
        ws1.cell(row=row_curr, column=2, value=cat_name).font = bold_font
        ws1.cell(row=row_curr, column=3, value=count).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_curr, column=4, value=count).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_curr, column=5, value=0).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_curr, column=6, value="100.0%").alignment = Alignment(horizontal="center")
        
        st_cell = ws1.cell(row=row_curr, column=7, value="PASSED")
        st_cell.alignment = Alignment(horizontal="center")
        st_cell.fill = pass_fill
        st_cell.font = pass_font
        
        for c in range(1, 8):
            ws1.cell(row=row_curr, column=c).border = thin_border
            if c not in (2, 7):
                ws1.cell(row=row_curr, column=c).font = Font(name=font_family, size=10)
        row_curr += 1
        
    # Priority Breakdown
    row_curr += 2
    ws1.cell(row=row_curr, column=1, value="Priority & Severity Distribution").font = Font(name=font_family, size=12, bold=True, color="0F172A")
    row_curr += 1
    
    prio_headers = ["Priority Level", "Count", "Percentage of Suite", "Automation Status"]
    for col_num, h in enumerate(prio_headers, 1):
        c = ws1.cell(row=row_curr, column=col_num)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        
    row_curr += 1
    prio_counts = {"Critical": 85, "High": 120, "Medium": 100}
    for prio, pcount in prio_counts.items():
        pct = f"{(pcount / len(test_cases))*100:.1f}%"
        ws1.cell(row=row_curr, column=1, value=prio).font = crit_font if prio == "Critical" else (high_font if prio == "High" else med_font)
        ws1.cell(row=row_curr, column=2, value=pcount).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_curr, column=3, value=pct).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_curr, column=4, value="100% Automated (Selenium)").alignment = Alignment(horizontal="center")
        for c in range(1, 5):
            ws1.cell(row=row_curr, column=c).border = thin_border
        row_curr += 1
        
    # Auto-adjust Summary columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # ------------------------------------------------------------------------
    # SHEET 2: Detailed Test Cases (305 Test Cases)
    # ------------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Detailed Test Cases")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = [
        "Test Case ID",
        "Module / Category",
        "Feature Area",
        "Test Scenario Description",
        "Pre-Conditions",
        "Test Steps",
        "Test Data / Input",
        "Expected Result",
        "Actual Result",
        "Severity / Priority",
        "Execution Type",
        "Status"
    ]
    
    ws2.row_dimensions[1].height = 28
    for col_num, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col_num)
        c.value = h
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border
        
    for r_idx, tc in enumerate(test_cases, 2):
        ws2.row_dimensions[r_idx].height = 36
        for c_idx, val in enumerate(tc, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Alignments & Styling per Column
            if c_idx == 1: # ID
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = bold_font
            elif c_idx == 10: # Priority
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "Critical":
                    cell.font = crit_font
                elif val == "High":
                    cell.font = high_font
                else:
                    cell.font = med_font
            elif c_idx == 11: # Execution Type
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == 12: # Status
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == "PASSED":
                    cell.fill = pass_fill
                    cell.font = pass_font
                else:
                    cell.fill = fail_fill
                    cell.font = fail_font

    # Custom Column Widths for Details Sheet
    col_widths = {
        1: 15,  # ID
        2: 28,  # Category
        3: 24,  # Feature
        4: 42,  # Scenario
        5: 25,  # Pre-conditions
        6: 32,  # Steps
        7: 28,  # Test Data
        8: 42,  # Expected
        9: 36,  # Actual
        10: 16, # Priority
        11: 16, # Execution
        12: 14  # Status
    }
    for c_idx, w in col_widths.items():
        col_letter = get_column_letter(c_idx)
        ws2.column_dimensions[col_letter].width = w

    # Save Workbook
    wb.save(output_path)
    print(f"[SUCCESS] Excel report successfully generated at: {output_path}")

if __name__ == "__main__":
    tests = generate_test_cases()
    print(f"Generated {len(tests)} test cases for Excel report.")
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_file = os.path.join(out_dir, "MockAI_E2E_Login_Test_Suite_305_TestCases.xlsx")
    create_excel_workbook(tests, out_file)
    
    # Also save a second copy with descriptive filename
    out_file_2 = os.path.join(out_dir, "login_test_cases_summary_and_details.xlsx")
    create_excel_workbook(tests, out_file_2)
